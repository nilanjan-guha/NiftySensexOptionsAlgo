# import os
# from datetime import datetime, timedelta

# import numpy as np
# import pandas as pd
# from dotenv import load_dotenv
# from kiteconnect import KiteConnect

# from strategy import (
#     calculate_indicators,
#     find_nifty_setups,
#     find_premium_setups,
#     get_setup_strength,
#     recommended_rr,
#     calculate_dynamic_risk,
#     PULLBACK_BUFFER,
#     MAX_PULLBACK_CANDLES,
#     MAX_SETUP_GAP_MINUTES,
# )


# # =========================================================
# # CONFIG
# # =========================================================

# NIFTY_INDEX_TOKEN = 256265

# STRIKE_INTERVAL = 50

# LOT_SIZE = 65
# LOTS = 1
# QUANTITY = LOT_SIZE * LOTS

# BACKTEST_DAYS = 100

# MARKET_START = "09:15"
# MARKET_END = "15:30"

# # =========================================================
# # DAILY TRADE MANAGEMENT
# # =========================================================
# #
# # Maximum 3 trades.
# #
# # Trade 1 WIN  -> STOP
# # Trade 1 LOSS -> Trade 2
# #
# # Trade 2 WIN  -> STOP
# # Trade 2 LOSS -> Trade 3
# #
# # Trade 3       -> STOP regardless of result
# #
# # This is risk control only.
# # It does NOT guarantee profitability.
# # =========================================================

# MAX_TRADES_PER_DAY = 3

# EXCEL_FILE = os.path.join(
#     os.path.dirname(
#         os.path.abspath(__file__)
#     ),
#     "Nifty_Options_Backtest.xlsx",
# )


# # =========================================================
# # ZERODHA
# # =========================================================

# SCRIPT_DIR = os.path.dirname(
#     os.path.abspath(__file__)
# )

# load_dotenv()

# API_KEY = os.getenv(
#     "KITE_API_KEY"
# )

# if not API_KEY:

#     print(
#         "ERROR: KITE_API_KEY not found in .env"
#     )

#     raise SystemExit(1)


# try:

#     with open(
#         os.path.join(
#             SCRIPT_DIR,
#             "access_token.txt"
#         ),
#         "r",
#         encoding="utf-8",
#     ) as f:

#         ACCESS_TOKEN = f.read().strip()

# except Exception:

#     print(
#         "ERROR: access_token.txt not found"
#     )

#     raise SystemExit(1)


# if not ACCESS_TOKEN:

#     print(
#         "ERROR: access_token.txt is empty"
#     )

#     raise SystemExit(1)


# kite = KiteConnect(
#     api_key=API_KEY
# )

# kite.set_access_token(
#     ACCESS_TOKEN
# )


# # =========================================================
# # EXCEL SAFE
# # =========================================================

# def make_excel_safe(df):

#     result = df.copy()

#     for col in result.columns:

#         if pd.api.types.is_datetime64_any_dtype(
#             result[col]
#         ):

#             if (
#                 getattr(
#                     result[col].dt,
#                     "tz",
#                     None
#                 )
#                 is not None
#             ):

#                 result[col] = (
#                     result[col]
#                     .dt
#                     .tz_localize(None)
#                 )

#         elif result[col].dtype == object:

#             def clean(x):

#                 if isinstance(
#                     x,
#                     pd.Timestamp
#                 ):

#                     return (
#                         x.tz_localize(None)
#                         if x.tzinfo
#                         else x
#                     )

#                 return x

#             result[col] = result[
#                 col
#             ].map(clean)

#     return result


# # =========================================================
# # HISTORICAL DATA
# # =========================================================

# def fetch_historical(
#     token,
#     start_date,
#     end_date,
# ):

#     try:

#         return kite.historical_data(
#             instrument_token=token,
#             from_date=start_date,
#             to_date=end_date,
#             interval="5minute",
#         )

#     except Exception as exc:

#         print(
#             "ZERODHA ERROR:",
#             type(exc).__name__,
#             str(exc),
#         )

#         return []


# def fetch_historical_chunked(
#     token,
#     start_date,
#     end_date,
# ):

#     all_candles = []

#     cursor = start_date

#     while cursor < end_date:

#         chunk_end = min(
#             cursor + timedelta(days=99),
#             end_date,
#         )

#         print(
#             "  Fetch:",
#             cursor,
#             "->",
#             chunk_end,
#         )

#         candles = fetch_historical(
#             token,
#             cursor,
#             chunk_end,
#         )

#         if candles:

#             all_candles.extend(
#                 candles
#             )

#         cursor = (
#             chunk_end
#             + timedelta(seconds=1)
#         )

#     if not all_candles:
#         return []

#     seen = {}

#     for candle in all_candles:

#         seen[
#             str(candle.get("date"))
#         ] = candle

#     return list(
#         seen.values()
#     )


# # =========================================================
# # DATE RANGE
# # =========================================================

# def get_backtest_range():

#     end_date = datetime.now()

#     start_date = (
#         end_date
#         - timedelta(
#             days=BACKTEST_DAYS
#         )
#     )

#     return (
#         start_date,
#         end_date
#     )


# def get_trading_days(df):

#     if df.empty:
#         return []

#     return sorted(
#         df["date"]
#         .dt
#         .date
#         .drop_duplicates()
#         .tolist()
#     )


# # =========================================================
# # HISTORICAL OPTION
# # =========================================================

# def find_option_for_date(
#     instruments,
#     trade_date,
#     direction,
#     spot,
# ):

#     atm = round(
#         spot / STRIKE_INTERVAL
#     ) * STRIKE_INTERVAL

#     candidates = []

#     for instrument in instruments:

#         try:

#             if instrument.get(
#                 "name"
#             ) != "NIFTY":
#                 continue

#             if instrument.get(
#                 "instrument_type"
#             ) != direction:
#                 continue

#             if float(
#                 instrument.get(
#                     "strike",
#                     0
#                 )
#             ) != float(atm):
#                 continue

#             expiry = instrument.get(
#                 "expiry"
#             )

#             if not expiry:
#                 continue

#             if expiry < trade_date:
#                 continue

#             candidates.append(
#                 instrument
#             )

#         except Exception:

#             continue

#     if not candidates:
#         return None

#     candidates.sort(
#         key=lambda x: x["expiry"]
#     )

#     return candidates[0]


# # =========================================================
# # TRADE EVALUATION
# # =========================================================

# def evaluate_trade(
#     premium_df,
#     entry_index,
#     entry_price,
#     stop_loss,
#     target,
# ):

#     for i in range(
#         entry_index + 1,
#         len(premium_df)
#     ):

#         candle = premium_df.iloc[i]

#         high = float(
#             candle["high"]
#         )

#         low = float(
#             candle["low"]
#         )

#         candle_time = (
#             candle["date"]
#         )

#         # Conservative:
#         # if target + SL happen in same candle,
#         # count SL first.
#         if (
#             high >= target
#             and low <= stop_loss
#         ):

#             return {
#                 "result": "LOSS",
#                 "exit_price": stop_loss,
#                 "exit_time": candle_time,
#                 "reason":
#                     "TARGET_AND_SL_SAME_CANDLE",
#             }

#         if high >= target:

#             return {
#                 "result": "WIN",
#                 "exit_price": target,
#                 "exit_time": candle_time,
#                 "reason": "TARGET",
#             }

#         if low <= stop_loss:

#             return {
#                 "result": "LOSS",
#                 "exit_price": stop_loss,
#                 "exit_time": candle_time,
#                 "reason": "STOP_LOSS",
#             }

#     last = premium_df.iloc[-1]

#     return {
#         "result": "OPEN",
#         "exit_price": float(
#             last["close"]
#         ),
#         "exit_time": last["date"],
#         "reason": "END_OF_DATA",
#     }


# # =========================================================
# # MAIN
# # =========================================================

# print(
#     "\n=========================================="
# )

# print(
#     "      EMA20 + VWAP ONLY BACKTEST"
# )

# print(
#     "=========================================="
# )

# print("\nConfiguration")

# print(
#     "Backtest days      :",
#     BACKTEST_DAYS
# )

# print(
#     "Timeframe          : 5 minutes"
# )

# print(
#     "Indicators         : EMA20 + VWAP ONLY"
# )

# print(
#     "Lot size            :",
#     LOT_SIZE
# )

# print(
#     "Lots                :",
#     LOTS
# )

# print(
#     "Quantity/trade      :",
#     QUANTITY
# )

# print(
#     "Max trades/day     :",
#     MAX_TRADES_PER_DAY
# )

# print(
#     "Entry               : Premium pullback CLOSE"
# )

# print(
#     "R:R                 : 1:1 / 1:1.5 / 1:2"
# )

# print(
#     "Excel               :",
#     EXCEL_FILE
# )


# start_date, end_date = (
#     get_backtest_range()
# )

# print(
#     "\nBacktest From:",
#     start_date
# )

# print(
#     "Backtest To  :",
#     end_date
# )


# # =========================================================
# # NIFTY
# # =========================================================

# print(
#     "\nFetching NIFTY historical data..."
# )

# nifty_raw = fetch_historical_chunked(
#     NIFTY_INDEX_TOKEN,
#     start_date,
#     end_date,
# )

# if not nifty_raw:

#     print(
#         "No NIFTY historical data."
#     )

#     raise SystemExit(1)


# nifty_df = calculate_indicators(
#     nifty_raw
# )

# if nifty_df.empty:

#     print(
#         "NIFTY indicator calculation failed."
#     )

#     raise SystemExit(1)


# trading_days = get_trading_days(
#     nifty_df
# )

# print(
#     "NIFTY candles :",
#     len(nifty_df)
# )

# print(
#     "Trading days  :",
#     len(trading_days)
# )


# # =========================================================
# # NFO
# # =========================================================

# print(
#     "\nDownloading NFO instruments..."
# )

# try:

#     instruments = kite.instruments(
#         "NFO"
#     )

# except Exception as exc:

#     print(
#         "Could not download NFO instruments."
#     )

#     print(
#         str(exc)
#     )

#     raise SystemExit(1)


# print(
#     "NFO instruments:",
#     len(instruments)
# )


# # =========================================================
# # BACKTEST
# # =========================================================

# results = []

# nifty_setup_count = 0

# premium_setup_count = 0


# for day_no, trade_date in enumerate(
#     trading_days,
#     start=1,
# ):

#     print(
#         f"\n[{day_no}/{len(trading_days)}] {trade_date}"
#     )

#     day_df = nifty_df[
#         nifty_df["date"].dt.date
#         == trade_date
#     ].copy().reset_index(
#         drop=True
#     )

#     if len(day_df) < 3:
#         continue

#     nifty_setups = (
#         find_nifty_setups(
#             day_df
#         )
#     )

#     if not nifty_setups:

#         print(
#             "  NIFTY setup: NONE"
#         )

#         continue

#     nifty_setup_count += len(
#         nifty_setups
#     )

#     trades_today = 0

#     last_trade_exit_time = None


#     # =====================================================
#     # PROCESS NIFTY SETUPS
#     # =====================================================

#     for nifty_setup in nifty_setups:

#         # Daily risk rule.
#         if (
#             trades_today
#             >= MAX_TRADES_PER_DAY
#         ):
#             break


#         direction = (
#             nifty_setup["direction"]
#         )

#         nifty_breakout = (
#             nifty_setup["breakout"]
#         )

#         nifty_pullback = (
#             nifty_setup["pullback"]
#         )


#         # Do not overlap trades.
#         if (
#             last_trade_exit_time
#             is not None
#             and
#             pd.Timestamp(
#                 nifty_breakout["date"]
#             )
#             <= last_trade_exit_time
#         ):
#             continue


#         spot = float(
#             nifty_pullback["close"]
#         )


#         # =================================================
#         # FIND OPTION
#         # =================================================

#         option = find_option_for_date(
#             instruments,
#             trade_date,
#             direction,
#             spot,
#         )

#         if option is None:

#             print(
#                 "  Historical option: NOT FOUND"
#             )

#             continue


#         option_token = (
#             option["instrument_token"]
#         )

#         option_symbol = (
#             option["tradingsymbol"]
#         )


#         # =================================================
#         # OPTION MARKET HOURS
#         # =================================================

#         day_start = datetime.combine(
#             trade_date,
#             datetime.strptime(
#                 MARKET_START,
#                 "%H:%M",
#             ).time(),
#         )

#         day_end = datetime.combine(
#             trade_date,
#             datetime.strptime(
#                 MARKET_END,
#                 "%H:%M",
#             ).time(),
#         )


#         premium_raw = fetch_historical(
#             option_token,
#             day_start,
#             day_end,
#         )

#         if not premium_raw:
#             continue


#         premium_df = (
#             calculate_indicators(
#                 premium_raw
#             )
#         )

#         if premium_df.empty:
#             continue


#         # =================================================
#         # PREMIUM SETUP
#         # =================================================

#         premium_setups = (
#             find_premium_setups(
#                 premium_df,
#                 pd.Timestamp(
#                     nifty_breakout["date"]
#                 ),
#                 pd.Timestamp(
#                     nifty_pullback["date"]
#                 ),
#             )
#         )

#         if not premium_setups:
#             continue


#         premium_setup = (
#             premium_setups[0]
#         )

#         premium_setup_count += 1


#         premium_breakout = (
#             premium_setup["breakout"]
#         )

#         premium_pullback = (
#             premium_setup["pullback"]
#         )


#         # =================================================
#         # ENTRY
#         # =================================================
#         #
#         # IMPORTANT:
#         #
#         # The pullback candle itself is CLOSED.
#         #
#         # Entry = pullback candle CLOSE.
#         #
#         # No future confirmation candle.
#         # =================================================

#         entry_time = pd.Timestamp(
#             premium_pullback["date"]
#         )

#         entry_price = float(
#             premium_pullback["close"]
#         )


#         matching = premium_df.index[
#             premium_df["date"]
#             == entry_time
#         ].tolist()

#         if not matching:
#             continue


#         entry_index = matching[0]


#         # =================================================
#         # STRENGTH
#         # =================================================

#         strength = get_setup_strength(
#             nifty_breakout,
#             nifty_pullback,
#             premium_breakout,
#             premium_pullback,
#         )


#         rr = recommended_rr(
#             strength
#         )

#         if rr is None:
#             continue


#         # =================================================
#         # RISK
#         # =================================================

#         risk = calculate_dynamic_risk(
#             entry_price,
#             premium_pullback,
#             premium_pullback,
#             direction,
#             strength,
#         )

#         if risk is None:
#             continue


#         # =================================================
#         # RESULT
#         # =================================================

#         trade_result = evaluate_trade(
#             premium_df,
#             entry_index,
#             risk["entry"],
#             risk["stop_loss"],
#             risk["target"],
#         )


#         pnl_points = (
#             trade_result["exit_price"]
#             - risk["entry"]
#         )

#         pnl_rupees = (
#             pnl_points
#             * QUANTITY
#         )


#         result = {

#             "date":
#                 trade_date,

#             "trade_number":
#                 trades_today + 1,

#             "direction":
#                 direction,

#             "option":
#                 option_symbol,

#             "expiry":
#                 option["expiry"],

#             "nifty_breakout":
#                 nifty_breakout["date"],

#             "nifty_pullback":
#                 nifty_pullback["date"],

#             "premium_breakout":
#                 premium_breakout["date"],

#             "premium_pullback":
#                 premium_pullback["date"],

#             "setup_strength":
#                 strength,

#             "selected_rr":
#                 f"1:{rr}",

#             "entry":
#                 risk["entry"],

#             "stop_loss":
#                 risk["stop_loss"],

#             "risk_points":
#                 risk["risk_points"],

#             "target":
#                 risk["target"],

#             "result":
#                 trade_result["result"],

#             "exit":
#                 trade_result["exit_price"],

#             "exit_time":
#                 trade_result["exit_time"],

#             "reason":
#                 trade_result["reason"],

#             "pnl_points":
#                 pnl_points,

#             "pnl":
#                 pnl_rupees,

#             "quantity":
#                 QUANTITY,
#         }


#         results.append(
#             result
#         )

#         trades_today += 1


#         if (
#             trade_result["result"]
#             != "OPEN"
#         ):

#             last_trade_exit_time = (
#                 pd.Timestamp(
#                     trade_result[
#                         "exit_time"
#                     ]
#                 )
#             )


#         print(
#             "  Trade",
#             trades_today,
#             "|",
#             direction,
#             "|",
#             option_symbol,
#             "| RR",
#             f"1:{rr}",
#             "|",
#             strength,
#             "|",
#             trade_result["result"],
#             "| P&L",
#             round(
#                 pnl_rupees,
#                 2
#             ),
#         )


#         # =================================================
#         # DAILY STOP RULE
#         # =================================================
#         #
#         # WIN -> STOP
#         # LOSS -> next trade
#         #
#         # Maximum 3 trades.
#         # =================================================

#         if (
#             trade_result["result"]
#             == "WIN"
#         ):

#             print(
#                 "  Daily WIN -> STOP TRADING"
#             )

#             break


#         if (
#             trade_result["result"]
#             == "LOSS"
#             and
#             trades_today
#             >= MAX_TRADES_PER_DAY
#         ):

#             print(
#                 "  Maximum daily trades reached"
#             )

#             break


# # =========================================================
# # RESULTS DATAFRAME
# # =========================================================

# results_df = pd.DataFrame(
#     results
# )


# if results_df.empty:

#     results_df = pd.DataFrame(
#         columns=[
#             "date",
#             "trade_number",
#             "direction",
#             "option",
#             "expiry",
#             "nifty_breakout",
#             "nifty_pullback",
#             "premium_breakout",
#             "premium_pullback",
#             "setup_strength",
#             "selected_rr",
#             "entry",
#             "stop_loss",
#             "risk_points",
#             "target",
#             "result",
#             "exit",
#             "exit_time",
#             "reason",
#             "pnl_points",
#             "pnl",
#             "quantity",
#         ]
#     )


# # =========================================================
# # SUMMARY
# # =========================================================

# wins = int(
#     (
#         results_df["result"]
#         == "WIN"
#     ).sum()
# )

# losses = int(
#     (
#         results_df["result"]
#         == "LOSS"
#     ).sum()
# )

# open_trades = int(
#     (
#         results_df["result"]
#         == "OPEN"
#     ).sum()
# )

# closed = (
#     wins + losses
# )

# win_rate = (
#     wins / closed * 100
#     if closed
#     else 0
# )


# gross_profit = (
#     results_df.loc[
#         results_df["pnl"] > 0,
#         "pnl",
#     ].sum()
# )

# gross_loss = (
#     results_df.loc[
#         results_df["pnl"] < 0,
#         "pnl",
#     ].sum()
# )

# net_pnl = (
#     results_df["pnl"].sum()
# )


# profit_factor = (
#     gross_profit
#     / abs(gross_loss)
#     if gross_loss < 0
#     else np.inf
# )


# average_trade = (
#     results_df["pnl"].mean()
#     if not results_df.empty
#     else 0
# )


# best_trade = (
#     results_df["pnl"].max()
#     if not results_df.empty
#     else 0
# )


# worst_trade = (
#     results_df["pnl"].min()
#     if not results_df.empty
#     else 0
# )


# # =========================================================
# # CE / PE SUMMARY
# # =========================================================

# direction_rows = []

# for direction in [
#     "CE",
#     "PE"
# ]:

#     d = results_df[
#         results_df["direction"]
#         == direction
#     ]

#     w = int(
#         (
#             d["result"]
#             == "WIN"
#         ).sum()
#     )

#     l = int(
#         (
#             d["result"]
#             == "LOSS"
#         ).sum()
#     )

#     c = w + l

#     direction_rows.append({

#         "Direction":
#             direction,

#         "Trades":
#             len(d),

#         "Wins":
#             w,

#         "Losses":
#             l,

#         "Open":
#             int(
#                 (
#                     d["result"]
#                     == "OPEN"
#                 ).sum()
#             ),

#         "Win Rate %":
#             round(
#                 w / c * 100,
#                 2
#             )
#             if c
#             else 0,

#         "Gross Profit":
#             d.loc[
#                 d["pnl"] > 0,
#                 "pnl"
#             ].sum(),

#         "Gross Loss":
#             d.loc[
#                 d["pnl"] < 0,
#                 "pnl"
#             ].sum(),

#         "Net P&L":
#             d["pnl"].sum(),
#     })


# direction_df = pd.DataFrame(
#     direction_rows
# )


# # =========================================================
# # TRADE NUMBER SUMMARY
# # =========================================================

# trade_number_rows = []

# for number in [
#     1,
#     2,
#     3,
# ]:

#     d = results_df[
#         results_df[
#             "trade_number"
#         ]
#         == number
#     ]

#     w = int(
#         (
#             d["result"]
#             == "WIN"
#         ).sum()
#     )

#     l = int(
#         (
#             d["result"]
#             == "LOSS"
#         ).sum()
#     )

#     c = w + l

#     trade_number_rows.append({

#         "Trade Number":
#             number,

#         "Trades":
#             len(d),

#         "Wins":
#             w,

#         "Losses":
#             l,

#         "Win Rate %":
#             round(
#                 w / c * 100,
#                 2
#             )
#             if c
#             else 0,

#         "Net P&L":
#             d["pnl"].sum(),
#     })


# trade_number_df = pd.DataFrame(
#     trade_number_rows
# )


# # =========================================================
# # MONTHLY SUMMARY
# # =========================================================

# monthly_rows = []

# if not results_df.empty:

#     results_df["month"] = (
#         pd.to_datetime(
#             results_df["date"]
#         )
#         .dt
#         .to_period("M")
#         .astype(str)
#     )

#     for month, d in (
#         results_df
#         .groupby("month")
#     ):

#         w = int(
#             (
#                 d["result"]
#                 == "WIN"
#             ).sum()
#         )

#         l = int(
#             (
#                 d["result"]
#                 == "LOSS"
#             ).sum()
#         )

#         c = w + l

#         monthly_rows.append({

#             "Month":
#                 month,

#             "Trades":
#                 len(d),

#             "Wins":
#                 w,

#             "Losses":
#                 l,

#             "Open":
#                 int(
#                     (
#                         d["result"]
#                         == "OPEN"
#                     ).sum()
#                 ),

#             "Win Rate %":
#                 round(
#                     w / c * 100,
#                     2
#                 )
#                 if c
#                 else 0,

#             "Profit":
#                 d.loc[
#                     d["pnl"] > 0,
#                     "pnl"
#                 ].sum(),

#             "Loss":
#                 d.loc[
#                     d["pnl"] < 0,
#                     "pnl"
#                 ].sum(),

#             "Net P&L":
#                 d["pnl"].sum(),
#         })


# monthly_df = pd.DataFrame(
#     monthly_rows
# )


# # =========================================================
# # SUMMARY DATA
# # =========================================================

# summary_data = {

#     "Backtest Start":
#         start_date,

#     "Backtest End":
#         end_date,

#     "Backtest Days":
#         BACKTEST_DAYS,

#     "Trading Days":
#         len(trading_days),

#     "NIFTY Setups":
#         nifty_setup_count,

#     "Premium Setups":
#         premium_setup_count,

#     "Completed Trades":
#         len(results_df),

#     "Wins":
#         wins,

#     "Losses":
#         losses,

#     "Open Trades":
#         open_trades,

#     "Closed Trades":
#         closed,

#     "Win Rate %":
#         round(
#             win_rate,
#             2
#         ),

#     "Gross Profit":
#         gross_profit,

#     "Gross Loss":
#         gross_loss,

#     "Net P&L":
#         net_pnl,

#     "Profit Factor":
#         (
#             round(
#                 profit_factor,
#                 2
#             )
#             if np.isfinite(
#                 profit_factor
#             )
#             else "INF"
#         ),

#     "Average Trade P&L":
#         average_trade,

#     "Best Trade":
#         best_trade,

#     "Worst Trade":
#         worst_trade,

#     "Lot Size":
#         LOT_SIZE,

#     "Lots":
#         LOTS,

#     "Quantity":
#         QUANTITY,
# }


# # =========================================================
# # SETTINGS
# # =========================================================

# settings_df = pd.DataFrame({

#     "Setting": [

#         "Backtest Days",

#         "Timeframe",

#         "Indicators",

#         "Lot Size",

#         "Lots",

#         "Quantity",

#         "Max Trades Per Day",

#         "Pullback Buffer",

#         "Max Pullback Candles",

#         "Max Setup Gap Minutes",

#         "R:R",

#         "Entry Rule",

#         "Daily Rule",

#     ],

#     "Value": [

#         BACKTEST_DAYS,

#         "5 minutes",

#         "EMA20 + VWAP ONLY",

#         LOT_SIZE,

#         LOTS,

#         QUANTITY,

#         MAX_TRADES_PER_DAY,

#         PULLBACK_BUFFER,

#         MAX_PULLBACK_CANDLES,

#         MAX_SETUP_GAP_MINUTES,

#         "1:1 / 1:1.5 / 1:2",

#         "Premium pullback candle CLOSE",

#         "WIN = STOP, LOSS = next trade, max 3",

#     ],
# })


# # =========================================================
# # EXCEL
# # =========================================================

# def create_excel():

#     try:

#         from openpyxl import (
#             load_workbook
#         )

#         from openpyxl.styles import (
#             Font,
#             PatternFill,
#             Alignment,
#         )

#         from openpyxl.utils import (
#             get_column_letter
#         )

#     except ImportError:

#         print(
#             "\nERROR: openpyxl is not installed."
#         )

#         print(
#             "Run: pip install openpyxl"
#         )

#         raise SystemExit(1)


#     frames = {

#         "Trade Details":
#             make_excel_safe(
#                 results_df
#             ),

#         "Summary":
#             pd.DataFrame(
#                 list(
#                     summary_data.items()
#                 ),
#                 columns=[
#                     "Metric",
#                     "Value"
#                 ],
#             ),

#         "CE PE Summary":
#             make_excel_safe(
#                 direction_df
#             ),

#         "Trade Number":
#             make_excel_safe(
#                 trade_number_df
#             ),

#         "Monthly Summary":
#             make_excel_safe(
#                 monthly_df
#             ),

#         "Settings":
#             make_excel_safe(
#                 settings_df
#             ),
#     }


#     with pd.ExcelWriter(
#         EXCEL_FILE,
#         engine="openpyxl",
#         mode="w",
#     ) as writer:

#         for sheet, frame in frames.items():

#             frame.to_excel(
#                 writer,
#                 sheet_name=sheet,
#                 index=False,
#             )


#     wb = load_workbook(
#         EXCEL_FILE
#     )


#     header_fill = PatternFill(
#         fill_type="solid",
#         fgColor="1F4E78",
#     )

#     header_font = Font(
#         bold=True,
#         color="FFFFFF",
#     )


#     for ws in wb.worksheets:

#         ws.freeze_panes = "A2"

#         if (
#             ws.max_row > 1
#             and ws.max_column > 0
#         ):

#             ws.auto_filter.ref = (
#                 ws.dimensions
#             )


#         for cell in ws[1]:

#             cell.fill = (
#                 header_fill
#             )

#             cell.font = (
#                 header_font
#             )

#             cell.alignment = (
#                 Alignment(
#                     horizontal="center",
#                     vertical="center",
#                 )
#             )


#         for column_cells in ws.columns:

#             idx = (
#                 column_cells[0]
#                 .column
#             )

#             max_len = 0

#             for cell in column_cells:

#                 max_len = max(
#                     max_len,
#                     len(
#                         str(
#                             cell.value
#                         )
#                     ),
#                 )

#             ws.column_dimensions[
#                 get_column_letter(
#                     idx
#                 )
#             ].width = min(
#                 max(
#                     max_len + 2,
#                     12
#                 ),
#                 45,
#             )


#         for row in ws.iter_rows():

#             for cell in row:

#                 if isinstance(
#                     cell.value,
#                     float
#                 ):

#                     cell.number_format = (
#                         "0.00"
#                     )


#     # =====================================================
#     # WIN / LOSS HIGHLIGHT
#     # =====================================================

#     if (
#         "Trade Details"
#         in wb.sheetnames
#     ):

#         ws = wb[
#             "Trade Details"
#         ]

#         headers = {

#             c.value:
#                 c.column

#             for c in ws[1]
#         }


#         if "result" in headers:

#             col = headers[
#                 "result"
#             ]

#             for row in range(
#                 2,
#                 ws.max_row + 1
#             ):

#                 value = ws.cell(
#                     row,
#                     col
#                 ).value

#                 if value == "WIN":

#                     ws.cell(
#                         row,
#                         col
#                     ).fill = PatternFill(
#                         fill_type="solid",
#                         fgColor="C6EFCE",
#                     )

#                 elif value == "LOSS":

#                     ws.cell(
#                         row,
#                         col
#                     ).fill = PatternFill(
#                         fill_type="solid",
#                         fgColor="FFC7CE",
#                     )


#     wb.save(
#         EXCEL_FILE
#     )


# # =========================================================
# # FINAL RESULT
# # =========================================================

# print(
#     "\n=========================================="
# )

# print(
#     "             BACKTEST RESULT"
# )

# print(
#     "          EMA20 + VWAP ONLY"
# )

# print(
#     "=========================================="
# )

# print(
#     "Trading days    :",
#     len(trading_days)
# )

# print(
#     "NIFTY setups    :",
#     nifty_setup_count
# )

# print(
#     "Premium setups  :",
#     premium_setup_count
# )

# print(
#     "Trades          :",
#     len(results_df)
# )

# print(
#     "Wins            :",
#     wins
# )

# print(
#     "Losses          :",
#     losses
# )

# print(
#     "Win rate        :",
#     round(
#         win_rate,
#         2
#     ),
#     "%"
# )

# print(
#     "Gross profit    :",
#     round(
#         gross_profit,
#         2
#     )
# )

# print(
#     "Gross loss      :",
#     round(
#         gross_loss,
#         2
#     )
# )

# print(
#     "Net P&L         :",
#     round(
#         net_pnl,
#         2
#     )
# )

# print(
#     "Profit factor   :",
#     (
#         round(
#             profit_factor,
#             2
#         )
#         if np.isfinite(
#             profit_factor
#         )
#         else "INF"
#     )
# )


# print(
#     "\nCE / PE"
# )

# print(
#     direction_df.to_string(
#         index=False
#     )
# )


# print(
#     "\nTRADE NUMBER"
# )

# print(
#     trade_number_df.to_string(
#         index=False
#     )
# )


# print(
#     "\nCreating / updating Excel..."
# )

# create_excel()


# print(
#     "\nExcel created successfully:"
# )

# print(
#     EXCEL_FILE
# )

# print(
#     "\nBacktest complete."
# )





import os
import time
from datetime import datetime, timedelta

import numpy as np
import pandas as pd
from dotenv import load_dotenv
from kiteconnect import KiteConnect

from strategy import (
    calculate_indicators,
    find_nifty_setups,
    find_sensex_setups,
    find_premium_setups,
    find_sensex_premium_setups,
    get_setup_strength,
    calculate_dynamic_risk,
)


# =========================================================
# CONFIG
# =========================================================

MARKETS = {

    "NIFTY": {

        "index_token": 256265,

        "spot_symbol": "NSE:NIFTY 50",

        "exchange": "NFO",

        "option_exchange": "NFO",

        "underlying_name": "NIFTY",

        "strike_interval": 50,

        "default_lot": 65,
    },

    "SENSEX": {

        "index_token": 265,

        "spot_symbol": "BSE:SENSEX",

        "exchange": "BFO",

        "option_exchange": "BFO",

        "underlying_name": "SENSEX",

        "strike_interval": 100,

        "default_lot": 20,
    },
}


BACKTEST_DAYS = 100

MARKET_START = "09:15"
MARKET_END = "15:30"

MAX_TRADES_PER_DAY_PER_MARKET = 3

LOTS = 1


# =========================================================
# ZERODHA RATE LIMIT
# =========================================================

# Zerodha historical API allows approximately
# 3 requests per second.
#
# We intentionally stay below that.
# 0.40 sec = approximately 2.5 requests/sec.

HISTORICAL_REQUEST_DELAY = 0.40

LAST_HISTORICAL_REQUEST = 0.0


# =========================================================
# FILE
# =========================================================

EXCEL_FILE = os.path.join(
    os.path.dirname(
        os.path.abspath(__file__)
    ),
    "Nifty_Sensex_Options_Backtest.xlsx",
)


# =========================================================
# ZERODHA
# =========================================================

SCRIPT_DIR = os.path.dirname(
    os.path.abspath(__file__)
)

load_dotenv()

API_KEY = os.getenv(
    "KITE_API_KEY"
)

if not API_KEY:

    print(
        "ERROR: KITE_API_KEY not found."
    )

    raise SystemExit(1)


try:

    with open(
        os.path.join(
            SCRIPT_DIR,
            "access_token.txt"
        ),
        "r",
        encoding="utf-8"
    ) as f:

        ACCESS_TOKEN = f.read().strip()

except Exception:

    print(
        "ERROR: access_token.txt not found."
    )

    raise SystemExit(1)


kite = KiteConnect(
    api_key=API_KEY
)

kite.set_access_token(
    ACCESS_TOKEN
)


# =========================================================
# CACHES
# =========================================================

# Exact historical request cache.
HISTORICAL_CACHE = {}


# Option data cache.
#
# Key:
#     (option_token, trade_date)
#
# This prevents the same option/day from
# being downloaded more than once.
OPTION_DAY_CACHE = {}


# =========================================================
# RATE LIMITED HISTORICAL REQUEST
# =========================================================

def rate_limited_historical(
    token,
    start_date,
    end_date,
    interval="5minute",
):
    """
    Safe historical-data wrapper.

    Features:
        1. Request throttling
        2. Exact-request caching
        3. Error handling
    """

    global LAST_HISTORICAL_REQUEST

    cache_key = (
        int(token),
        str(start_date),
        str(end_date),
        interval,
    )

    # -----------------------------------------------------
    # CACHE
    # -----------------------------------------------------

    if cache_key in HISTORICAL_CACHE:

        return HISTORICAL_CACHE[
            cache_key
        ]


    # -----------------------------------------------------
    # RATE LIMIT
    # -----------------------------------------------------

    elapsed = (
        time.monotonic()
        -
        LAST_HISTORICAL_REQUEST
    )

    if elapsed < HISTORICAL_REQUEST_DELAY:

        time.sleep(
            HISTORICAL_REQUEST_DELAY
            -
            elapsed
        )


    # -----------------------------------------------------
    # REQUEST
    # -----------------------------------------------------

    try:

        candles = kite.historical_data(

            instrument_token=int(
                token
            ),

            from_date=start_date,

            to_date=end_date,

            interval=interval,
        )

        LAST_HISTORICAL_REQUEST = (
            time.monotonic()
        )


        if candles is None:

            candles = []


        HISTORICAL_CACHE[
            cache_key
        ] = candles


        return candles


    except Exception as exc:

        LAST_HISTORICAL_REQUEST = (
            time.monotonic()
        )

        print()

        print(
            "ZERODHA HISTORICAL ERROR"
        )

        print(
            "Token:",
            token
        )

        print(
            "From:",
            start_date
        )

        print(
            "To:",
            end_date
        )

        print(
            "Error:",
            type(exc).__name__,
            str(exc)
        )

        print()

        HISTORICAL_CACHE[
            cache_key
        ] = []

        return []


# =========================================================
# EXCEL SAFE
# =========================================================

def make_excel_safe(df):

    result = df.copy()

    if result.empty:

        return result


    for col in result.columns:

        if pd.api.types.is_datetime64_any_dtype(
            result[col]
        ):

            try:

                if result[col].dt.tz is not None:

                    result[col] = (
                        result[col]
                        .dt
                        .tz_localize(None)
                    )

            except Exception:

                pass


        elif result[col].dtype == object:

            def clean(value):

                if isinstance(
                    value,
                    pd.Timestamp
                ):

                    if value.tzinfo:

                        return value.tz_localize(
                            None
                        )

                return value


            result[col] = (
                result[col]
                .map(clean)
            )


    return result


# =========================================================
# HISTORICAL
# =========================================================

def fetch_historical(
    token,
    start_date,
    end_date
):

    return rate_limited_historical(
        token,
        start_date,
        end_date,
        "5minute",
    )


# =========================================================
# CHUNKED HISTORICAL
# =========================================================

def fetch_chunked(
    token,
    start_date,
    end_date
):

    all_candles = []

    cursor = start_date


    while cursor < end_date:

        chunk_end = min(
            cursor + timedelta(days=89),
            end_date
        )


        candles = fetch_historical(
            token,
            cursor,
            chunk_end
        )


        if candles:

            all_candles.extend(
                candles
            )


        cursor = (
            chunk_end
            +
            timedelta(seconds=1)
        )


    unique = {}


    for candle in all_candles:

        unique[
            str(candle["date"])
        ] = candle


    return list(
        unique.values()
    )


# =========================================================
# DATE RANGE
# =========================================================

def get_backtest_range():

    end = datetime.now()

    start = (
        end
        -
        timedelta(
            days=BACKTEST_DAYS
        )
    )

    return start, end


# =========================================================
# TRADING DAYS
# =========================================================

def get_trading_days(df):

    if df.empty:

        return []


    return sorted(
        df["date"]
        .dt.date
        .drop_duplicates()
        .tolist()
    )


# =========================================================
# FIND OPTION
# =========================================================

def find_option_for_date(
    instruments,
    market_config,
    trade_date,
    direction,
    spot
):

    atm = (
        round(
            spot
            /
            market_config[
                "strike_interval"
            ]
        )
        *
        market_config[
            "strike_interval"
        ]
    )


    candidates = []


    for instrument in instruments:

        try:

            if instrument.get(
                "name"
            ) != market_config[
                "underlying_name"
            ]:

                continue


            if instrument.get(
                "instrument_type"
            ) != direction:

                continue


            if float(
                instrument.get(
                    "strike",
                    0
                )
            ) != float(atm):

                continue


            expiry = instrument.get(
                "expiry"
            )


            if not expiry:

                continue


            if expiry < trade_date:

                continue


            candidates.append(
                instrument
            )


        except Exception:

            continue


    if not candidates:

        return None


    candidates.sort(
        key=lambda x:
        x["expiry"]
    )


    return candidates[0]


# =========================================================
# OPTION DAY DATA
# =========================================================

def get_option_day_data(
    option_token,
    trade_date,
):
    """
    Fetch option 5-minute candles for one day.

    The same token/date combination
    will only be requested once.
    """

    cache_key = (
        int(option_token),
        trade_date,
    )


    if cache_key in OPTION_DAY_CACHE:

        return OPTION_DAY_CACHE[
            cache_key
        ]


    day_start = datetime.combine(
        trade_date,
        datetime.strptime(
            MARKET_START,
            "%H:%M"
        ).time()
    )


    day_end = datetime.combine(
        trade_date,
        datetime.strptime(
            MARKET_END,
            "%H:%M"
        ).time()
    )


    raw = fetch_historical(
        option_token,
        day_start,
        day_end
    )


    if not raw:

        OPTION_DAY_CACHE[
            cache_key
        ] = pd.DataFrame()

        return OPTION_DAY_CACHE[
            cache_key
        ]


    df = calculate_indicators(
        raw
    )


    OPTION_DAY_CACHE[
        cache_key
    ] = df


    return df


# =========================================================
# TRADE EVALUATION
# =========================================================

def evaluate_trade(
    premium_df,
    entry_index,
    entry,
    stop_loss,
    target
):

    for i in range(
        entry_index + 1,
        len(premium_df)
    ):

        candle = premium_df.iloc[i]


        high = float(
            candle["high"]
        )


        low = float(
            candle["low"]
        )


        # -------------------------------------------------
        # BOTH TARGET AND SL SAME CANDLE
        # -------------------------------------------------

        if (
            high >= target
            and
            low <= stop_loss
        ):

            return {

                "result":
                    "LOSS",

                "exit_price":
                    stop_loss,

                "exit_time":
                    candle["date"],

                "reason":
                    "TARGET_AND_SL_SAME_CANDLE",
            }


        # -------------------------------------------------
        # TARGET
        # -------------------------------------------------

        if high >= target:

            return {

                "result":
                    "WIN",

                "exit_price":
                    target,

                "exit_time":
                    candle["date"],

                "reason":
                    "TARGET",
            }


        # -------------------------------------------------
        # STOP LOSS
        # -------------------------------------------------

        if low <= stop_loss:

            return {

                "result":
                    "LOSS",

                "exit_price":
                    stop_loss,

                "exit_time":
                    candle["date"],

                "reason":
                    "STOP_LOSS",
            }


    # -----------------------------------------------------
    # END OF DATA
    # -----------------------------------------------------

    last = premium_df.iloc[-1]


    return {

        "result":
            "OPEN",

        "exit_price":
            float(
                last["close"]
            ),

        "exit_time":
            last["date"],

        "reason":
            "END_OF_DATA",
    }


# =========================================================
# MAIN
# =========================================================

print()

print(
    "=========================================="
)

print(
    " NIFTY + SENSEX OPTIONS BACKTEST"
)

print(
    "      EMA20 + VWAP ONLY"
)

print(
    "=========================================="
)

print(
    "Backtest days:",
    BACKTEST_DAYS
)

print(
    "Lot: NIFTY 65 / SENSEX runtime lot"
)

print(
    "R:R: Dynamic based on setup strength"
)

print()


# =========================================================
# DATE RANGE
# =========================================================

start_date, end_date = (
    get_backtest_range()
)


# =========================================================
# DOWNLOAD NFO + BFO INSTRUMENTS
# =========================================================

instrument_cache = {}


for market_name, config in MARKETS.items():

    print()

    print(
        "Downloading",
        config[
            "option_exchange"
        ],
        "instruments..."
    )


    try:

        instrument_cache[
            market_name
        ] = kite.instruments(
            config[
                "option_exchange"
            ]
        )


        print(
            market_name,
            "instruments:",
            len(
                instrument_cache[
                    market_name
                ]
            )
        )


    except Exception as exc:

        print(
            market_name,
            "instrument error:",
            str(exc)
        )


        instrument_cache[
            market_name
        ] = []


# =========================================================
# RESULTS
# =========================================================

results = []


# =========================================================
# MARKET BY MARKET
# =========================================================

for market_name, config in MARKETS.items():

    print()

    print(
        "=========================================="
    )

    print(
        "BACKTESTING:",
        market_name
    )

    print(
        "=========================================="
    )

    print(
        "Index token:",
        config[
            "index_token"
        ]
    )

    print(
        "Option exchange:",
        config[
            "option_exchange"
        ]
    )


    # =====================================================
    # INDEX DATA
    # =====================================================

    print()

    print(
        "Fetching",
        market_name,
        "index data..."
    )


    raw = fetch_chunked(
        config[
            "index_token"
        ],
        start_date,
        end_date
    )


    if not raw:

        print(
            "No",
            market_name,
            "historical data."
        )

        continue


    index_df = calculate_indicators(
        raw
    )


    if index_df.empty:

        continue


    trading_days = (
        get_trading_days(
            index_df
        )
    )


    print(
        "Trading days:",
        len(trading_days)
    )


    # =====================================================
    # DAY BY DAY
    # =====================================================

    for day_number, trade_date in enumerate(
        trading_days,
        start=1
    ):

        print(
            f"[{market_name}] "
            f"Day {day_number}/"
            f"{len(trading_days)} "
            f"- {trade_date}",
            flush=True
        )


        day_df = index_df[
            index_df["date"].dt.date
            ==
            trade_date
        ].copy()


        day_df = (
            day_df
            .reset_index(drop=True)
        )


        if len(day_df) < 2:

            continue


        # =================================================
        # UNDERLYING SETUP
        # =================================================

        if market_name == "NIFTY":

            setups = find_nifty_setups(
                day_df
            )

        else:

            setups = find_sensex_setups(
                day_df
            )


        if not setups:

            continue


        trades_today = 0


        # =================================================
        # EACH SETUP
        # =================================================

        for setup in setups:

            if (
                trades_today
                >=
                MAX_TRADES_PER_DAY_PER_MARKET
            ):

                break


            direction = setup[
                "direction"
            ]


            breakout = setup[
                "breakout"
            ]


            pullback = setup[
                "pullback"
            ]


            # =================================================
            # SPOT REFERENCE
            # =================================================

            spot = float(
                pullback[
                    "close"
                ]
            )


            # =================================================
            # FIND ATM OPTION
            # =================================================

            option = (
                find_option_for_date(
                    instrument_cache[
                        market_name
                    ],
                    config,
                    trade_date,
                    direction,
                    spot
                )
            )


            if option is None:

                continue


            option_token = option[
                "instrument_token"
            ]


            option_symbol = option[
                "tradingsymbol"
            ]


            option_lot = int(
                option.get(
                    "lot_size",
                    config[
                        "default_lot"
                    ]
                )
            )


            quantity = (
                option_lot
                *
                LOTS
            )


            # =================================================
            # OPTION DATA
            # =================================================

            premium_df = (
                get_option_day_data(
                    option_token,
                    trade_date
                )
            )


            if premium_df.empty:

                continue


            # =================================================
            # PREMIUM SETUP
            # =================================================

            if market_name == "NIFTY":

                premium_setups = (
                    find_premium_setups(
                        premium_df,
                        pd.Timestamp(
                            breakout[
                                "date"
                            ]
                        ),
                        pd.Timestamp(
                            pullback[
                                "date"
                            ]
                        )
                    )
                )

            else:

                premium_setups = (
                    find_sensex_premium_setups(
                        premium_df,
                        pd.Timestamp(
                            breakout[
                                "date"
                            ]
                        ),
                        pd.Timestamp(
                            pullback[
                                "date"
                            ]
                        )
                    )
                )


            if not premium_setups:

                continue


            premium_setup = (
                premium_setups[0]
            )


            premium_breakout = (
                premium_setup[
                    "breakout"
                ]
            )


            premium_pullback = (
                premium_setup[
                    "pullback"
                ]
            )


            premium_confirmation = (
                premium_setup.get(
                    "confirmation"
                )
            )


            # =================================================
            # ENTRY
            # =================================================

            entry_time = pd.Timestamp(
                premium_pullback[
                    "date"
                ]
            )


            entry = float(
                premium_pullback[
                    "close"
                ]
            )


            matching = premium_df.index[
                premium_df["date"]
                ==
                entry_time
            ].tolist()


            if not matching:

                continue


            entry_index = matching[0]


            # =================================================
            # SETUP STRENGTH
            # =================================================

            setup_strength = (
                get_setup_strength(
                    pullback,
                    pullback,
                    premium_breakout,
                    premium_pullback,
                    premium_confirmation,
                )
            )


            # =================================================
            # RISK
            # =================================================

            risk = (
                calculate_dynamic_risk(
                    entry,
                    premium_pullback,
                    premium_confirmation,
                    direction,
                    setup_strength,
                )
            )


            if risk is None:

                continue


            # =================================================
            # RESULT
            # =================================================

            trade_result = (
                evaluate_trade(
                    premium_df,
                    entry_index,
                    risk[
                        "entry"
                    ],
                    risk[
                        "stop_loss"
                    ],
                    risk[
                        "target"
                    ]
                )
            )


            pnl_points = (
                trade_result[
                    "exit_price"
                ]
                -
                risk[
                    "entry"
                ]
            )


            pnl = (
                pnl_points
                *
                quantity
            )


            # =================================================
            # STORE TRADE
            # =================================================

            results.append({

                "market":
                    market_name,

                "date":
                    trade_date,

                "direction":
                    direction,

                "option":
                    option_symbol,

                "exchange":
                    config[
                        "option_exchange"
                    ],

                "expiry":
                    option[
                        "expiry"
                    ],

                "lot_size":
                    option_lot,

                "quantity":
                    quantity,

                "underlying_breakout":
                    breakout[
                        "date"
                    ],

                "underlying_pullback":
                    pullback[
                        "date"
                    ],

                "premium_breakout":
                    premium_breakout[
                        "date"
                    ],

                "premium_pullback":
                    premium_pullback[
                        "date"
                    ],

                "premium_confirmation":
                    (
                        premium_confirmation[
                            "date"
                        ]
                        if premium_confirmation
                        is not None
                        else None
                    ),

                "setup_strength":
                    setup_strength,

                "entry":
                    risk[
                        "entry"
                    ],

                "stop_loss":
                    risk[
                        "stop_loss"
                    ],

                "risk_points":
                    risk[
                        "risk_points"
                    ],

                "rr":
                    risk[
                        "rr"
                    ],

                "target":
                    risk[
                        "target"
                    ],

                "result":
                    trade_result[
                        "result"
                    ],

                "exit":
                    trade_result[
                        "exit_price"
                    ],

                "exit_time":
                    trade_result[
                        "exit_time"
                    ],

                "reason":
                    trade_result[
                        "reason"
                    ],

                "pnl_points":
                    pnl_points,

                "pnl":
                    pnl,

            })


            trades_today += 1


# =========================================================
# RESULTS
# =========================================================

results_df = pd.DataFrame(
    results
)


if results_df.empty:

    print()

    print(
        "No trades found."
    )

    print()

    raise SystemExit(0)


# =========================================================
# MARKET SUMMARY
# =========================================================

market_rows = []


for market_name in MARKETS.keys():

    d = results_df[
        results_df["market"]
        ==
        market_name
    ]


    wins = int(
        (
            d["result"]
            ==
            "WIN"
        ).sum()
    )


    losses = int(
        (
            d["result"]
            ==
            "LOSS"
        ).sum()
    )


    closed = (
        wins
        +
        losses
    )


    gross_profit = (
        d.loc[
            d["pnl"] > 0,
            "pnl"
        ].sum()
    )


    gross_loss = (
        d.loc[
            d["pnl"] < 0,
            "pnl"
        ].sum()
    )


    net = d[
        "pnl"
    ].sum()


    pf = (
        gross_profit
        /
        abs(gross_loss)
        if gross_loss < 0
        else np.inf
    )


    market_rows.append({

        "Market":
            market_name,

        "Trades":
            len(d),

        "Wins":
            wins,

        "Losses":
            losses,

        "Win Rate %":
            round(
                wins
                /
                closed
                *
                100,
                2
            )
            if closed
            else 0,

        "Gross Profit":
            gross_profit,

        "Gross Loss":
            gross_loss,

        "Net P&L":
            net,

        "Profit Factor":
            round(
                pf,
                2
            )
            if np.isfinite(pf)
            else "INF",
    })


market_summary_df = pd.DataFrame(
    market_rows
)


# =========================================================
# OVERALL
# =========================================================

wins = int(
    (
        results_df["result"]
        ==
        "WIN"
    ).sum()
)


losses = int(
    (
        results_df["result"]
        ==
        "LOSS"
    ).sum()
)


closed = (
    wins
    +
    losses
)


gross_profit = (
    results_df.loc[
        results_df["pnl"] > 0,
        "pnl"
    ].sum()
)


gross_loss = (
    results_df.loc[
        results_df["pnl"] < 0,
        "pnl"
    ].sum()
)


net_pnl = (
    results_df[
        "pnl"
    ].sum()
)


profit_factor = (
    gross_profit
    /
    abs(gross_loss)
    if gross_loss < 0
    else np.inf
)


win_rate = (
    wins
    /
    closed
    *
    100
    if closed
    else 0
)


# =========================================================
# SUMMARY
# =========================================================

summary_df = pd.DataFrame({

    "Metric": [

        "Backtest Start",

        "Backtest End",

        "Backtest Days",

        "Markets",

        "Overall Trades",

        "Wins",

        "Losses",

        "Win Rate %",

        "Gross Profit",

        "Gross Loss",

        "Net P&L",

        "Profit Factor",

        "NIFTY Lot",

        "SENSEX Lot",

        "Strategy",
    ],


    "Value": [

        start_date,

        end_date,

        BACKTEST_DAYS,

        "NIFTY + SENSEX",

        len(results_df),

        wins,

        losses,

        round(
            win_rate,
            2
        ),

        gross_profit,

        gross_loss,

        net_pnl,

        (
            round(
                profit_factor,
                2
            )
            if np.isfinite(
                profit_factor
            )
            else "INF"
        ),

        MARKETS[
            "NIFTY"
        ][
            "default_lot"
        ],

        MARKETS[
            "SENSEX"
        ][
            "default_lot"
        ],

        "EMA20 + VWAP ONLY",
    ]
})


# =========================================================
# EXCEL
# =========================================================

def create_excel():

    from openpyxl import load_workbook

    from openpyxl.styles import (
        Font,
        PatternFill,
        Alignment
    )

    from openpyxl.utils import (
        get_column_letter
    )


    frames = {

        "Trade Details":
            make_excel_safe(
                results_df
            ),

        "Market Summary":
            make_excel_safe(
                market_summary_df
            ),

        "Overall Summary":
            make_excel_safe(
                summary_df
            ),
    }


    with pd.ExcelWriter(
        EXCEL_FILE,
        engine="openpyxl",
        mode="w"
    ) as writer:

        for sheet, frame in (
            frames.items()
        ):

            frame.to_excel(
                writer,
                sheet_name=sheet,
                index=False
            )


    wb = load_workbook(
        EXCEL_FILE
    )


    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78"
    )


    header_font = Font(
        bold=True,
        color="FFFFFF"
    )


    for ws in wb.worksheets:

        ws.freeze_panes = "A2"

        ws.auto_filter.ref = (
            ws.dimensions
        )


        for cell in ws[1]:

            cell.fill = (
                header_fill
            )

            cell.font = (
                header_font
            )

            cell.alignment = (
                Alignment(
                    horizontal="center",
                    vertical="center"
                )
            )


        for column_cells in ws.columns:

            idx = (
                column_cells[0]
                .column
            )


            max_len = max(
                len(
                    str(
                        cell.value
                    )
                )
                for cell
                in column_cells
            )


            ws.column_dimensions[
                get_column_letter(
                    idx
                )
            ].width = min(
                max(
                    max_len + 2,
                    12
                ),
                45
            )


    wb.save(
        EXCEL_FILE
    )


# =========================================================
# FINAL PRINT
# =========================================================

print()

print(
    "=========================================="
)

print(
    "        NIFTY + SENSEX BACKTEST"
)

print(
    "          EMA20 + VWAP ONLY"
)

print(
    "=========================================="
)


print(
    market_summary_df.to_string(
        index=False
    )
)


print()

print(
    "OVERALL"
)


print(
    "Trades:",
    len(results_df)
)


print(
    "Wins:",
    wins
)


print(
    "Losses:",
    losses
)


print(
    "Win rate:",
    round(
        win_rate,
        2
    ),
    "%"
)


print(
    "Gross profit:",
    round(
        gross_profit,
        2
    )
)


print(
    "Gross loss:",
    round(
        gross_loss,
        2
    )
)


print(
    "Net P&L:",
    round(
        net_pnl,
        2
    )
)


print(
    "Profit factor:",
    (
        round(
            profit_factor,
            2
        )
        if np.isfinite(
            profit_factor
        )
        else "INF"
    )
)


print()

print(
    "Historical requests cached:",
    len(
        HISTORICAL_CACHE
    )
)


print(
    "Option day cache entries:",
    len(
        OPTION_DAY_CACHE
    )
)


print()

print(
    "Creating Excel..."
)


create_excel()


print()

print(
    "Excel created:"
)


print(
    EXCEL_FILE
)